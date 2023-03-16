
import os
import csv

####### DELETE FILE ######
def delete_view():
    folder_path = os.environ["USERPROFILE"]+'\Downloads'
    for file in os.listdir(folder_path):
        if(file.startswith("view")):
            filepath  = os.path.join(folder_path,file)
            os.remove(filepath)



asin_mp_language = {
    "FR" : 5,
    "GB" : 3,
    "DE" : 4,
    "US" : 1,
    "AU" : 111172,
    "IN" : 44571,
    "IT" : 35691,
    "ES" : 44551,
    "CA" : 7
    }


######## PARSE THE FILE #######
def collate_data(asin_mp_dict):
    folder_path = os.environ["USERPROFILE"]+'\Downloads'
    print(asin_mp_dict)
    dict_of_views = {}

    def create_dict(key,value,type_1, cs_name):
        dict_of_views[key] = {
            "value" : value,
            "type_1" : type_1,
            "cs_name" : cs_name
        }

    def check_row(data,header_object):
        
        if (len(data) == 0):
                return False
        #if(data.count("incontinence_protector_size") > 0) or (data.count("^incontinence_protector_size$") > 0):
                #check for value and customer 
        #if(header.find("value") and header.("customer_name")):
        if header_object.get('value') != None and header_object.get('customer_name') != None:
                    value =  data[header_object["value"]]
                    cs_name = data[header_object["customer_name"]]
                    item_id = data[header_object["item_id"]]
                    type_1 = data[header_object["type"]]
                    create_dict(item_id,value,type_1,cs_name)
                    return True
                    
        return False
            
                 



    for file in os.listdir(folder_path) :
        if file.startswith("view"):
            f = os.path.join(folder_path,file)
            print(f)
            with open(f, 'r') as file:
                csvreader = csv.reader(file)
                header_object = {}
                header = []
                body  = []
                index = 0 
                for row in csvreader:
                    if(index == 0):
                        header = row
                    else:
                        body.append(row)
                    index+=1
                print(header)
                for ind,head_value in enumerate(header):
                    if head_value.endswith("value"):
                        header_object["value"] = ind
                    elif head_value.endswith("customer_name"):
                        header_object["customer_name"] = ind
                    elif head_value.endswith("type") and len(head_value.split('_')) == 1:
                        header_object["type"] = ind     
                    elif head_value.endswith("language_tag"):
                        header_object["language_tag"] = ind
                    elif head_value.endswith("item_id"):
                        header_object["item_id"] = ind
                print(header_object)
                

                row_data = []
                for row in body:
                    print(row)
                    try:
                        language = row[header_object["language_tag"]].split("_") if row[header_object["language_tag"]].strip()!="" else []

                        print(language)
                        if(len(language) == 2):
                            mp_mapping  = asin_mp_language.get(language[1])
                            print(mp_mapping, asin_mp_dict[row[header_object["item_id"]]])
                            #if  mp_mapping != None:
                            if mp_mapping == asin_mp_dict[row[header_object["item_id"]]]:
                                    row_data = row
                                    print('----', asin_mp_dict[row[header_object["item_id"]]] )
                                    break
                    except:
                        pass

                print(row_data)

                if len(row_data) == 0 :
                    row_data = body[0]
                    create_dict(row_data[header_object["item_id"]],"No","No","No")
                else:
                    if check_row(row_data,header_object) == False:
                            print(header_object["item_id"] , "item")
                            create_dict(row_data[header_object["item_id"]],"No","No","No")


    print(dict_of_views)





### Read and Write to file

    import openpyxl
    folder_path = 'input.xlsx'

    workbook = openpyxl.load_workbook(filename=folder_path)
    sheet = workbook.active

    for row in range(2,sheet.max_row+1):
        asin = sheet.cell(row = row, column = 1)
        view_data = dict_of_views.get(asin.value)
        #print(asin.value)
        if view_data:
            for col in range(3,6):
                cell_obj = sheet.cell(row = row, column = col)
                if col == 3:
                    cell_obj.value = view_data.get("value") 
                elif col == 4:
                    cell_obj.value = view_data.get("type_1")
                else:
                    cell_obj.value = view_data.get("cs_name")  

    workbook.save(filename=folder_path)


